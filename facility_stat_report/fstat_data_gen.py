#/usr/bin/env python
# -*- coding: utf-8 -*-

# This file is an indirect dependancy for the script fstat_pdf_gen.py
# This creates an output file fstat_import_data.py which will have the
# data for the plots to be generated to be in the facility metrics pdf

import unicodedata

from openpyxl import load_workbook
from colour_science import SCILIFE_COLOURS, FACILITY_USER_AFFILIATION_COLOUR_OFFICIAL
from facility_report_data_gen import user_plot_data, publication_plot_data

def fix_spl_char(value):
    if value == None:
        value = ''
    elif not isinstance(value, unicode):
        value = unicode(str(value), 'utf-8')
    return str(unicodedata.normalize('NFKD', value).encode('ascii', 'ignore'))

fac_map = {'AIDA Data Hub' : ['AIDA Data Hub'],
            'BioImage Informatics' : ['BioImage Informatics'],
            'Compute and Storage' : ['Bioinformatics Compute and Storage'],
            'Support, Infrastructure and Training' : ['Bioinformatics Support, Infrastructure and Training', 'Bioinformatics Support and Infrastructure'],
            'Advanced Light Microscopy' : ['Advanced Light Microscopy (ALM)'],
            'Centre for Cellular Imaging' : ['Centre for Cellular Imaging'],
            'Intravital Microscopy Facility' : ['Intravital Microscopy Facility'],
            'Biochemical Imaging Centre Umeå' : ['Biochemical Imaging Centre Umeå'],
            'Cryo-EM' : ['Cryo-EM'],
            'Cell Profiling' : ['Cell Profiling'],
            'Advanced FISH Technologies' : ['Advanced FISH Technologies'],
            'In Situ Sequencing' : ['In Situ Sequencing (ISS)'],
            'National Resource for Mass Spectrometry Imaging' : ['National Resource for Mass Spectrometry Imaging'],
            'Gothenburg Imaging Mass Spectrometry' : ['Gothenburg Imaging Mass Spectrometry'],
            'Chemical Biology Consortium Sweden' : ['Chemical Biology Consortium Sweden (CBCS)'],
            'Chemical Proteomics' : ['Chemical Proteomics'],
            'Genome Engineering Zebrafish' : ['Genome Engineering Zebrafish'],
            'High Throughput Genome Engineering' : ['High-throughput Genome Engineering (HTGE)'],
            'Clinical Genomics Gothenburg' : ['Clinical Genomics Gothenburg'],
            'Clinical Genomics Linköping' : ['Clinical Genomics Linköping'],
            'Clinical Genomics Lund' : ['Clinical Genomics Lund'],
            'Clinical Genomics Stockholm' : ['Clinical Genomics Stockholm'],
            'Clinical Genomics Umeå' : ['Clinical Genomics Umeå'],
            'Clinical Genomics Uppsala' : ['Clinical Genomics Uppsala'],
            'Clinical Genomics Örebro' : ['Clinical Genomics Örebro'],
            'Drug Discovery and Development' : ['Drug Discovery and Development (DDD)'],
            'Ancient DNA' : ['Ancient DNA'],
            'Eukaryotic Single Cell Genomics' : ['Eukaryotic Single Cell Genomics (ESCG)'],
            'Microbial Single Cell Genomics' : ['Microbial Single Cell Genomics'],
            'National Genomics Infrastructure' : ['National Genomics Infrastructure'],
            'Autoimmunity Profiling' : ['Autoimmunity Profiling'],
            'Exposomics' : ['Exposomics'],
            'Glycoproteomics' : ['Glycoproteomics'],
            'Mass Cytometry' : ['Mass Cytometry'],
            'Proximity Proteomics' : ['PLA and Single Cell Proteomics'],
            'Plasma Profiling' : ['Plasma Profiling'],
            'Proteogenomics' : ['Proteogenomics'],
            'Targeted and Structural Proteomics' : ['Targeted and Structural Proteomics'],
            'Swedish Metabolomics Centre' : ['Swedish Metabolomics Centre (SMC)'],
            'Swedish NMR Centre' : ['Swedish NMR Centre (SNC)']}

pie_domains = {
    2017: "{x:[0, 0.30], y:[0, 0.5]}",
    2018: "{x:[0.35, 0.65], y:[0, 0.5]}",
    2019: "{x:[0.70, 1], y:[0, 0.5]}"
}

cat_col_ind = {"Service":0, "Collaborative":7, "Technology development":9, "None":5}
jif_col_ind = [0, 7, 9, 1, 5]

jif_names = ["JIF < 6", "JIF 6 - 9", "JIF 9 - 25", "JIF > 25", "JIF unknown"]

affiliation_data = {}
facility_data = {}

years_to_work = [2017, 2018, 2019]

for year in years_to_work:
    user_wb = load_workbook(filename="Infrastructure Users {}.xlsx".format(str(year)), read_only=True) #File provided by OO
    user_ws = user_wb["Users"]
    for rnum, row in enumerate(user_ws.rows):
        if rnum == 0:
            continue

        facility = row[1].value.strip()
        affiliation = row[4].value

        if not affiliation:
            print "WARN: Facility '{}' no affiliation, year {}, row {}".format(facility.encode('utf-8'), str(year), str(rnum))
            continue

        affiliation = affiliation.strip().replace("chalmers", "Chalmers")
        

        # check the reporting unit name are as desired
        if facility.encode('utf-8') not in fac_map.keys():
            print "WARN: Facility '{}' not known, year {}, row {}".format(facility.encode('utf-8'), str(year), str(rnum))
            continue
        
        facility = fix_spl_char(facility)

        # assign empty dict for facility
        if facility not in affiliation_data:
            affiliation_data[facility] = {}

        # assign empty dict for year
        if year not in affiliation_data[facility].keys():
            affiliation_data[facility][year] = {}

        if affiliation not in affiliation_data[facility][year].keys():
            affiliation_data[facility][year][affiliation] = 0

        affiliation_data[facility][year][affiliation] += 1

fstat_wb = load_workbook(filename="Facility Stat 2019.xlsx") #File provided by OO
fstat_ws = fstat_wb["Single Data"]

for rnum, row in enumerate(fstat_ws.rows):
    if rnum == 0:
        continue
    
    facility = row[1].value.strip()
    if facility.encode('utf-8') not in fac_map.keys():
        print "WARN: Facility '{}' not known for FSTAT, row {}".format(facility.encode('utf-8'), str(rnum))
        continue
    
    facility_data[fix_spl_char(facility)] = dict(
                                                fd = row[3].value.strip().encode('utf-8'),
                                                fh = row[4].value.strip().encode('utf-8') if row[4].value else 'N/A',
                                                sf = row[5].value,
                                                hu = row[6].value.strip().encode('utf-8'),
                                                fte = str(row[7].value).strip(),
                                                sfte = str(row[8].value).strip(),
                                                sfund = str(row[9].value).strip(),
                                                tfund = str(row[10].value).strip()
                                                )


pub_bar_data = {}

for fac in fac_map.keys():
# Create barplot for category
    pub_bar_data[fix_spl_char(fac)] = publication_plot_data(fac_map[fac], 2019)


py_tmp = '''#!/usr/bin/env python
# -*- coding: utf-8 -*-

affiliation_data = {afdata}

facility_data = {fcdata}

pub_bar_data = {pbdata}

fac_map = {fldata}

'''

with open("fstat_import_data.py", "w") as pfl:
    pfl.write(py_tmp.format(afdata=affiliation_data, fcdata=facility_data, pbdata=pub_bar_data, fldata=fac_map))
    
    